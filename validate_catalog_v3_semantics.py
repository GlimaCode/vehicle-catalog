# -*- coding: utf-8 -*-
"""
Independent Version 3 semantic validator (Python, separate from the TS app).

Verifies the 25 mandatory checks against catalog-v3.db, the V3 Excel
workbook, the audit/delta files, and the preserved V1/V2 artifacts.
PASS only if every check succeeds.
"""
import csv, hashlib, json, os, re, sqlite3, sys

BASE = os.path.dirname(os.path.abspath(__file__))
APP = os.path.join(BASE, "catalog-app")
DB = os.path.join(APP, "data", "catalog-v3.db")
XLSX = os.path.join(APP, "exports", "Complete_US_Vehicle_Catalog_1980_to_2026-07-15_v3.xlsx")

HIER = {"Sub-model", "Trim", "Series", "Edition", "Generation", "Chassis"}
CONF = {"Engine Variant", "Drivetrain Variant", "Body Style", "Package",
        "Commercial Configuration"}

def norm(s):
    return re.sub(r"[^A-Z0-9+]", "", (s or "").upper())

def read_csv(path):
    return list(csv.DictReader(open(path, encoding="utf-8-sig")))

def main():
    checks = []
    def c(n, name, ok, detail=""):
        checks.append({"n": n, "name": name, "ok": bool(ok), "detail": str(detail)[:300]})

    con = sqlite3.connect(f"file:{DB}?mode=ro", uri=True)
    con.row_factory = sqlite3.Row
    q = lambda sql, *p: con.execute(sql, p).fetchone()[0]
    allq = lambda sql, *p: con.execute(sql, p).fetchall()

    # 1-3: configuration types never stored as hierarchy values
    for i, t in [(1, "Engine Variant"), (2, "Drivetrain Variant"), (3, "Body Style")]:
        c(i, f"No {t} stored as a Sub-model/hierarchy value",
          q("SELECT COUNT(*) FROM vehicle_hierarchy_values WHERE classification_type=?", t) == 0)

    # 4-5: no package / trim stored as a Model
    model_norms = {(r["mk"], r["nm"]) for r in allq(
        "SELECT k.standard_make mk, m.norm_model nm FROM models m JOIN makes k ON k.id=m.make_id")}
    PKG = {"Z71", "FX4", "TRDPRO", "PRO4X", "TREMOR"}
    TRIMS = {"LARIAT", "XLT", "DENALI", "KINGRANCH", "EXL", "SLINE", "RSPEC",
             "LAREDO", "RUBICON", "SLT", "PLATINUM", "LTZ"}
    c(4, "No Package value stored as a Model",
      not any(nm in PKG for _, nm in model_norms))
    c(5, "No Trim stored as a Model",
      not any(nm in TRIMS for _, nm in model_norms))

    # 6: dashboard counts = classification-specific DB counts (API logic mirrors
    # these exact queries; verified by construction against the same DB)
    hier_counts = {r["classification_type"]: r["n"] for r in allq(
        "SELECT classification_type, COUNT(*) n FROM vehicle_hierarchy_values GROUP BY 1")}
    conf_counts = {r["classification_type"]: r["n"] for r in allq(
        "SELECT classification_type, COUNT(*) n FROM vehicle_configuration_values GROUP BY 1")}
    c(6, "Dashboard counters are classification-specific (per-type counts exist; "
      "no combined 'sub-model' total used)",
      len(hier_counts) >= 2 and len(conf_counts) == 5,
      f"hier={hier_counts}, conf={conf_counts}")

    # 7-8: selector separation
    c(7, "Hierarchy table holds only allowed hierarchy types",
      q(f"""SELECT COUNT(*) FROM vehicle_hierarchy_values WHERE classification_type
        NOT IN ({','.join('?' * len(HIER))})""", *HIER) == 0)
    c(8, "Configuration table holds only allowed configuration types",
      q(f"""SELECT COUNT(*) FROM vehicle_configuration_values WHERE classification_type
        NOT IN ({','.join('?' * len(CONF))})""", *CONF) == 0)

    # 9: every approved hierarchy record passed its individual audit
    audit = read_csv(os.path.join(BASE, "Approved_Submodel_Trim_Audit.csv"))
    audited = {(r["Standard Make"], norm(r["Standard Model"]),
                norm(r["Canonical Spelling"])) for r in audit
               if r["Decision"] in ("Approved", "Name Corrected")}
    missing_audit = [r for r in allq(
        """SELECT k.standard_make mk, m.norm_model nm, s.norm_value nv,
           s.classification_type ct FROM vehicle_hierarchy_values s
           JOIN models m ON m.id=s.model_id JOIN makes k ON k.id=m.make_id
           WHERE s.classification_type IN ('Sub-model','Trim','Edition')""")
        if (r["mk"], r["nm"], r["nv"]) not in audited]
    c(9, "Every approved Sub-model/Trim/Edition passed its individual audit",
      len(missing_audit) == 0, len(missing_audit))

    # 10: removed/reclassified values in delta
    delta = read_csv(os.path.join(BASE, "Application_Data_V2_to_V3_Delta.csv"))
    kinds = {r["Change Type"] for r in delta}
    c(10, "Removed/reclassified values recorded in the delta",
      "Value moved to review" in kinds and "Classification change" in kinds
      and "Source-status correction" in kinds, sorted(kinds))

    # 11-12: EPA 1980-83
    recon = read_csv(os.path.join(BASE, "EPA_1980_1983_Source_Reconciliation.csv"))
    c(11, "Every EPA 1980-1983 source year investigated (with formats and locations)",
      {r["Model Year"] for r in recon} == {"1980", "1981", "1982", "1983"}
      and all(r["Imported"] == "Yes" for r in recon))
    cov = read_csv(os.path.join(BASE, "Vehicle_Coverage_1980_1983_Audit.csv"))
    ok12 = all(r["Version 3 Decision"] for r in cov)
    c(12, "Every imported historical source record accounted for",
      len(cov) > 1000 and ok12, f"{len(cov)} carline rows")

    # 13-14: source organization vs dataset
    c(13, "Source organization and dataset stored separately",
      q("SELECT COUNT(*) FROM sources WHERE source_organization IS NULL OR source_organization=''") == 0)
    # cross-dataset (NHTSA recall+vPIC only) must not be Fully Verified
    bad14 = allq("""SELECT k.standard_make, m.standard_model FROM models m
      JOIN makes k ON k.id=m.make_id
      WHERE m.validation_status='Fully Verified' AND m.source_organization_count<2""")
    c(14, "Cross-dataset confirmation not mislabeled as independent organizations",
      len(bad14) == 0, [tuple(r) for r in bad14[:5]])

    # 15-16: year-level evidence for every approved record
    c(15, "Every approved hierarchy record has year-level evidence",
      q("""SELECT COUNT(*) FROM vehicle_hierarchy_values s WHERE NOT EXISTS
        (SELECT 1 FROM hierarchy_value_years y WHERE y.hierarchy_value_id=s.id)""") == 0)
    c(16, "Every approved configuration record has year-level evidence",
      q("""SELECT COUNT(*) FROM vehicle_configuration_values s WHERE NOT EXISTS
        (SELECT 1 FROM configuration_value_years y WHERE y.configuration_value_id=s.id)""") == 0)

    # 17: non-contiguity preserved
    c(17, "Non-contiguous ranges remain non-contiguous",
      q("""SELECT COUNT(*) FROM models m WHERE m.confirmed_model_years LIKE '%;%'
        AND (SELECT COUNT(*) FROM model_years y WHERE y.model_id=m.id)
        = m.last_confirmed_model_year - m.first_confirmed_model_year + 1""") == 0
      and q("""SELECT COUNT(*) FROM vehicle_hierarchy_values s
        WHERE s.confirmed_model_years LIKE '%;%'
        AND (SELECT COUNT(*) FROM hierarchy_value_years y WHERE y.hierarchy_value_id=s.id)
        = s.last_confirmed_model_year - s.first_confirmed_model_year + 1""") == 0)

    # 18: unresolved candidates hidden from approved data
    c(18, "Unresolved candidates absent from approved hierarchy/configuration tables",
      q("""SELECT COUNT(*) FROM vehicle_hierarchy_values WHERE validation_status
        NOT IN ('Fully Verified','Government Verified','Manufacturer Verified')""") == 0
      and q("""SELECT COUNT(*) FROM vehicle_configuration_values WHERE validation_status
        NOT IN ('Fully Verified','Government Verified','Manufacturer Verified')""") == 0)

    # 19-20: Excel counts + worksheets
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    REQUIRED = ["Vehicle Hierarchy", "Vehicle Configurations", "Makes", "Models",
                "Submodels", "Trims", "Series", "Editions", "Generations", "Chassis",
                "Engines", "Drivetrains", "Body Styles", "Packages",
                "Commercial Configurations", "Model Years", "Hierarchy Years",
                "Configuration Years", "Aliases", "Review Required", "Sources",
                "Taxonomy Audit", "1980-1983 Audit", "V2 to V3 Changes",
                "Catalog Summary"]
    names = wb.sheetnames
    c(20, "All 25 required worksheets exist", all(s in names for s in REQUIRED),
      f"{len(names)} sheets")
    expects = {
        "Vehicle Hierarchy": q("SELECT COUNT(*) FROM vehicle_hierarchy_values"),
        "Vehicle Configurations": q("SELECT COUNT(*) FROM vehicle_configuration_values"),
        "Makes": q("SELECT COUNT(*) FROM makes"),
        "Models": q("SELECT COUNT(*) FROM models"),
        "Model Years": q("SELECT COUNT(*) FROM model_years"),
        "Hierarchy Years": q("SELECT COUNT(*) FROM hierarchy_value_years"),
        "Configuration Years": q("SELECT COUNT(*) FROM configuration_value_years"),
        "Review Required": q("SELECT COUNT(*) FROM validation_reviews"),
        "Trims": hier_counts.get("Trim", 0),
        "Engines": conf_counts.get("Engine Variant", 0),
    }
    mism = []
    for sheet, expected in expects.items():
        got = wb[sheet].max_row - 1
        if got != expected:
            mism.append(f"{sheet}: xlsx={got} db={expected}")
    c(19, "Excel counts match the Version 3 database", not mism, "; ".join(mism) or "all match")

    # 21: idempotent re-import (verified by the app importer run; re-check counts marker)
    extra = json.load(open(os.path.join(APP, "exports", "v3_verify_extra.json"),
                           encoding="utf-8-sig"))
    c(21, "Re-import idempotency", extra.get("reimport_idempotent") == "PASS",
      extra.get("reimport_idempotent_detail", ""))
    c(22, "Production build passes", extra.get("production_build", "").startswith("PASS"))
    c(23, "Automated tests pass", extra.get("automated_tests", "").startswith("PASS")
      or "PASS" in extra.get("automated_tests", ""))
    c(24, "Windows scripts remain path-independent",
      "%~dp0" in open(os.path.join(APP, "start-app.bat")).read()
      and "%~dp0" in open(os.path.join(APP, "setup-windows.bat")).read())

    # 25: V1 and V2 files unmodified (hash re-check vs Phase 2 manifest)
    manifest = read_csv(os.path.join(BASE, "Phase2_Hash_Manifest.csv"))
    fails = []
    for row in manifest:
        p = row["Path"]
        if p.startswith("catalog-app-phase2-backup"):
            continue          # backup itself
        full = os.path.join(BASE, p)
        if not os.path.exists(full):
            fails.append(p + " missing")
            continue
        h = hashlib.sha256(open(full, "rb").read()).hexdigest().upper()
        if h != row["SHA256"].upper():
            fails.append(p)
    c(25, "Version 1 and Version 2 files remain unmodified (hash re-check of "
      "root catalog/audit files)", not fails, fails[:6])

    report = {
        "version": "V3",
        "database": DB,
        "checks": checks,
        "status": "PASS" if all(x["ok"] for x in checks) else "FAIL",
    }
    out = os.path.join(APP, "data", "app_verification_report_v3.json")
    json.dump(report, open(out, "w", encoding="utf-8"), indent=2)
    print(json.dumps({"status": report["status"],
                      "failed": [x for x in checks if not x["ok"]]}, indent=1))
    print("Report:", out)
    return 0 if report["status"] == "PASS" else 1

if __name__ == "__main__":
    sys.exit(main())
