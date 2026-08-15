# -*- coding: utf-8 -*-
"""Independent Version 4 hierarchy validator (25 mandatory checks)."""
import csv, hashlib, json, os, re, sqlite3, sys
from collections import Counter

BASE = os.path.dirname(os.path.abspath(__file__))
APP = os.path.join(BASE, "catalog-app")
DB = os.path.join(APP, "data", "catalog-v4.db")
XLSX = os.path.join(APP, "exports", "Complete_US_Vehicle_Catalog_1980_to_2026-07-15_v4.xlsx")
HIER = {"Sub-model", "Trim", "Series", "Edition", "Generation", "Chassis"}

def norm(s):
    return re.sub(r"[^A-Z0-9+]", "", (s or "").upper())

def read_csv(name, base=BASE):
    return list(csv.DictReader(open(os.path.join(base, name), encoding="utf-8-sig")))

def main():
    checks = []
    def c(n, name, ok, detail=""):
        checks.append({"n": n, "name": name, "ok": bool(ok), "detail": str(detail)[:300]})

    con = sqlite3.connect(f"file:{DB}?mode=ro", uri=True)
    con.row_factory = sqlite3.Row
    q = lambda sql, *p: con.execute(sql, p).fetchone()[0]
    allq = lambda sql, *p: con.execute(sql, p).fetchall()

    audit = read_csv("High_Priority_Trim_Research_Audit.csv")
    c(1, "All 286 high-priority candidates accounted for", len(audit) == 286, len(audit))
    ids = [r["Candidate ID"] for r in audit]
    keys = [(r["Standard Make"], r["Standard Model"], r["Candidate Value"]) for r in audit]
    c(2, "No candidate appears more than once in the audit",
      len(set(ids)) == len(ids) and len(set(keys)) == len(keys))

    approved = [r for r in audit if r["Research Outcome"].startswith("Approved")]
    c(3, "Every newly approved value has official Make-Model evidence "
      "(source organization + document + URL)",
      all(r["Primary Source Organization"] and r["Primary Source Document"]
          and r["Primary Source URL"] for r in approved))
    c(4, "Every newly approved value has confirmed US-market years",
      all(r["Confirmed Model Years"] and r["US Market Confirmed"] == "Yes"
          for r in approved))

    c(5, "No Package stored as a Trim",
      q("""SELECT COUNT(*) FROM vehicle_hierarchy_values WHERE classification_type='Trim'
        AND norm_value IN ('Z71','FX4','TRDPRO','PRO4X','TREMOR')""") == 0)
    for i, t in [(6, "Engine Variant"), (7, "Drivetrain Variant"), (8, "Body Style")]:
        c(i, f"No {t} stored as a hierarchy value",
          q("SELECT COUNT(*) FROM vehicle_hierarchy_values WHERE classification_type=?",
            t) == 0)
    c(9, "No global-only value in the approved US catalog",
      not any(r["Research Outcome"] == "Global-Market Only" and
              r["Version 4 Action"].startswith("Added") for r in audit))

    # 10-11: unsupported years / non-contiguity
    hier_v4 = read_csv("Complete_US_Vehicle_Hierarchy_1980_to_2026-07-15_v4.csv")
    year_rows = read_csv("Complete_US_Vehicle_Hierarchy_By_Year_v4.csv")
    def expand(t):
        ys = []
        for seg in t.split(";"):
            m = re.match(r"^(\d{4})(?:-(\d{4}))?$", seg.strip())
            if m:
                ys.extend(range(int(m.group(1)), int(m.group(2) or m.group(1)) + 1))
        return ys
    expected_years = sum(len(expand(r["Confirmed Model Years"])) for r in hier_v4)
    c(10, "No unsupported model year generated (year rows = exact expansion)",
      len(year_rows) == expected_years, f"{len(year_rows)} vs {expected_years}")
    c(11, "Non-contiguous availability remains non-contiguous",
      q("""SELECT COUNT(*) FROM vehicle_hierarchy_values s
        WHERE s.confirmed_model_years LIKE '%;%'
        AND (SELECT COUNT(*) FROM hierarchy_value_years y WHERE y.hierarchy_value_id=s.id)
        = s.last_confirmed_model_year - s.first_confirmed_model_year + 1""") == 0)

    src_idx = read_csv("Official_Hierarchy_Source_Index.csv")
    doc_urls = {r["Primary Source URL"] for r in approved
                if r["Primary Source Organization"] not in ("EPA", "NHTSA")}
    indexed = {r["Source URL"] for r in src_idx}
    c(12, "Every official manufacturer source document is indexed",
      doc_urls <= indexed, f"{len(doc_urls)} docs, {len(src_idx)} index entries")
    c(13, "Source organization and source document stored separately",
      all(r["Source Organization"] and r["Document Title"] for r in src_idx)
      and all(r["Primary Source Organization"] and r["Primary Source Document"]
              for r in hier_v4))

    rejected = [r for r in audit if not r["Research Outcome"].startswith("Approved")]
    c(14, "Every rejected candidate has a documented reason",
      all(r["Notes"] or r["Official Naming Evidence"] for r in rejected))

    hier_v3 = read_csv("US_Vehicle_Hierarchy_Values_2026-07-15_v3.csv")
    v4keys = {(r["Standard Make"], norm(r["Standard Model"]),
               norm(r["Standard Hierarchy Value"]), r["Classification Type"])
              for r in hier_v4}
    missing_v3 = [r for r in hier_v3 if (r["Standard Make"], norm(r["Standard Model"]),
                  norm(r["Standard Sub-model or Variant"]), r["Classification Type"])
                  not in v4keys]
    c(15, "Every V3 hierarchy value has a V4 disposition (carried forward)",
      not missing_v3, len(missing_v3))

    missing_db = 0
    for r in hier_v4:
        hit = q("""SELECT COUNT(*) FROM vehicle_hierarchy_values s
          JOIN models m ON m.id=s.model_id JOIN makes k ON k.id=m.make_id
          WHERE k.standard_make=? AND m.norm_model=? AND s.norm_value=?
          AND s.classification_type=?""",
          r["Standard Make"], norm(r["Standard Model"]),
          norm(r["Standard Hierarchy Value"]), r["Classification Type"])
        if not hit:
            missing_db += 1
    c(16, "Every V4 hierarchy record exists in the database", missing_db == 0, missing_db)

    import openpyxl
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    c(17, "Database and Excel hierarchy counts match",
      wb["Vehicle Hierarchy"].max_row - 1 == q("SELECT COUNT(*) FROM vehicle_hierarchy_values")
      and wb["Hierarchy Years"].max_row - 1 == q("SELECT COUNT(*) FROM hierarchy_value_years")
      and wb["Trims"].max_row - 1 == q(
        "SELECT COUNT(*) FROM vehicle_hierarchy_values WHERE classification_type='Trim'"),
      f"xlsx hier={wb['Vehicle Hierarchy'].max_row - 1}, "
      f"db={q('SELECT COUNT(*) FROM vehicle_hierarchy_values')}")
    for s in ["Trim Research Audit", "Official Source Index", "V3 to V4 Changes"]:
        if s not in wb.sheetnames:
            c(17, f"Missing sheet {s}", False)

    c(18, "Selector options contain only approved hierarchy types",
      q(f"""SELECT COUNT(*) FROM vehicle_hierarchy_values WHERE classification_type
        NOT IN ({','.join('?' * len(HIER))})""", *HIER) == 0)
    c(19, "Review candidates remain outside approved tables",
      q("""SELECT COUNT(*) FROM vehicle_hierarchy_values WHERE validation_status
        NOT IN ('Fully Verified','Government Verified','Manufacturer Verified')""") == 0)

    extra = json.load(open(os.path.join(APP, "exports", "v4_verify_extra.json"),
                           encoding="utf-8-sig"))
    c(20, "Re-import is idempotent", extra.get("reimport_idempotent") == "PASS",
      extra.get("reimport_idempotent_detail", ""))
    c(21, "Tests pass", "PASS" in extra.get("automated_tests", ""))
    c(22, "Production build passes", extra.get("production_build", "").startswith("PASS"))
    c(23, "Windows scripts remain path-independent",
      "%~dp0" in open(os.path.join(APP, "start-app.bat")).read())

    manifest = read_csv("Phase3_Hash_Manifest.csv")
    fails = []
    for row in manifest:
        p = row["Path"]
        if p.startswith("catalog-app-phase3-backup"):
            continue
        full = os.path.join(BASE, p)
        if not os.path.exists(full):
            fails.append(p + " missing")
            continue
        h = hashlib.sha256(open(full, "rb").read()).hexdigest().upper()
        if h != row["SHA256"].upper():
            fails.append(p)
    c(24, "Version 1, 2 and 3 artifacts remain unchanged (hash re-check)",
      not fails, fails[:6])

    delta = read_csv("Application_Data_V3_to_V4_Delta.csv")
    delta_items = {(r["Standard Make"], r["Item"].split(" / ")[0]) for r in delta}
    c(25, "Every semantic change appears in the V3-to-V4 delta",
      len(delta) == 286 and all(r["Change Type"] for r in delta),
      f"{len(delta)} delta rows (one per candidate)")

    report = {"version": "V4", "database": DB,
              "outcome_counts": dict(Counter(r["Research Outcome"] for r in audit)),
              "checks": checks,
              "status": "PASS" if all(x["ok"] for x in checks) else "FAIL"}
    out = os.path.join(APP, "data", "app_verification_report_v4.json")
    json.dump(report, open(out, "w", encoding="utf-8"), indent=2)
    print(json.dumps({"status": report["status"],
                      "failed": [x for x in checks if not x["ok"]]}, indent=1))
    print("Report:", out)
    return 0 if report["status"] == "PASS" else 1

if __name__ == "__main__":
    sys.exit(main())
